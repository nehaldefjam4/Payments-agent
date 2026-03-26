from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
green_fill = PatternFill("solid", fgColor="C6EFCE")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
num_fmt = '#,##0.00'
date_fmt = 'DD/MM/YYYY'

def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font, cell.border = data_font, thin_border
        if fill:
            cell.fill = fill

# ════════════════════════════════════════════════════════════════
# FILE 1: CBC_Master_Sheet.xlsx
# Contains ONLY old BOOKING transactions (Jan-Feb 2026).
# Q1/Q2 installments are NOT here — agent must discover + add them.
# ════════════════════════════════════════════════════════════════
wb = Workbook()

ws1 = wb.active
ws1.title = "Updated Sheet_Escrow Account"
headers_escrow = ["Date", "Value Date", "Narration", "Reference", "Debit", "Credit",
                  "Running Balance", "Unit No. / Remarks", "Account Name", "Receipt", "Status"]
for c, h in enumerate(headers_escrow, 1):
    ws1.cell(1, c, h)
style_header(ws1, 11)
for col, w in {'A':12,'B':12,'C':55,'D':22,'E':14,'F':14,'G':16,'H':18,'I':32,'J':10,'K':10}.items():
    ws1.column_dimensions[col].width = w

# 10 historical BOOKING payments — already reconciled
historical = [
    (datetime(2026,1,5), datetime(2026,1,5),
     "OUTWARD CLEARINGUNIT 101 DEFAULT - ~026ABC001", "REF-H001",
     0, 450000, "101", "Ahmed Hassan Al Maktoum", "Done", "NEW"),
    (datetime(2026,1,8), datetime(2026,1,8),
     "INWARD REMITTANCETT REF: 530P4F86D98E100 AED 380000 FATIMA ABDULRAHMAN SAEED PO BOX 5432 DUBAI",
     "REF-H002", 0, 380000, "102", "Fatima Abdulrahman Saeed", "Done", "NEW"),
    (datetime(2026,1,12), datetime(2026,1,12),
     "OUTWARD CLEARINGUNIT 203 DEFAULT - ~026ABC003", "REF-H003",
     0, 520000, "203", "Mohamed Khalid Al Rashidi", "Done", "NEW"),
    (datetime(2026,1,15), datetime(2026,1,15),
     "INWARD REMITTANCETT REF: 044ABLC7890100 AED 490000 SARAH JOHNSON 45 PARK LANE LONDON",
     "REF-H004", 0, 490000, "204", "Sarah Johnson", "Done", "NEW"),
    (datetime(2026,1,20), datetime(2026,1,20),
     "OUTWARD CLEARINGUNIT 305 DEFAULT - ~026ABC005", "REF-H005",
     0, 410000, "305", "Khalid Omar Al Mansouri", "Done", "NEW"),
    (datetime(2026,1,25), datetime(2026,1,25),
     "INWARD REMITTANCETT REF: 20260125001000 AED 375000 NADIA HASSAN IBRAHIM",
     "REF-H006", 0, 375000, "306", "Nadia Hassan Ibrahim", "Done", "NEW"),
    (datetime(2026,2,1), datetime(2026,2,1),
     "OUTWARD CLEARINGUNIT 407 DEFAULT - ~026ABC007", "REF-H007",
     0, 560000, "407", "Rashid Saeed Al Ketbi", "Done", "NEW"),
    (datetime(2026,2,5), datetime(2026,2,5),
     "INWARD REMITTANCETT REF: 530P5EB4FCD500 AED 430000 ELENA PETROVA MOSCOW RUSSIA",
     "REF-H008", 0, 430000, "408", "Elena Petrova", "Done", "NEW"),
    (datetime(2026,2,10), datetime(2026,2,10),
     "OUTWARD CLEARINGUNIT 509 DEFAULT - ~026ABC009", "REF-H009",
     0, 510000, "509", "Ali Mohammed Al Suwaidi", "Done", "NEW"),
    (datetime(2026,2,15), datetime(2026,2,15),
     "INWARD REMITTANCETT REF: 20260215005000 AED 395000 JAMES ROBERT WILSON FLAT 12 CHICAGO",
     "REF-H010", 0, 395000, "510", "James Robert Wilson", "Done", "NEW"),
]

balance = 5000000.00
for i, (dt, vd, narr, ref, deb, cred, unit, name, rcpt, status) in enumerate(historical):
    r = i + 2
    ws1.cell(r, 1, dt).number_format = date_fmt
    ws1.cell(r, 2, vd).number_format = date_fmt
    ws1.cell(r, 3, narr)
    ws1.cell(r, 4, ref)
    ws1.cell(r, 5, deb).number_format = num_fmt
    ws1.cell(r, 6, cred).number_format = num_fmt
    if i == 0:
        ws1.cell(r, 7, balance - deb + cred).number_format = num_fmt
    else:
        ws1.cell(r, 7, f"=G{r-1}-E{r}+F{r}")
    ws1.cell(r, 7).number_format = num_fmt
    ws1.cell(r, 8, unit)
    ws1.cell(r, 9, name)
    ws1.cell(r, 10, rcpt)
    ws1.cell(r, 11, status)
    style_row(ws1, r, 11, green_fill)

# Corporate sheet
ws2 = wb.create_sheet("Updated Sheet_Corporate")
headers_corp = ["Date", "Narration", "Reference", "Debit", "Credit", "Running Balance",
                "Unit No. / Remarks", "Account Name"]
for c, h in enumerate(headers_corp, 1):
    ws2.cell(1, c, h)
style_header(ws2, 8)
for col, w in {'A':12,'B':50,'C':20,'D':14,'E':14,'F':16,'G':16,'H':30}.items():
    ws2.column_dimensions[col].width = w

corp_data = [
    (datetime(2026,1,10), "SERVICE CHARGE - CBC COMMON AREAS Q1 2026", "CORP-001", 125000, 0, "CBC Management", ""),
    (datetime(2026,2,1), "DEWA UTILITIES PAYMENT - CBC BUILDING", "CORP-002", 45000, 0, "DEWA", ""),
    (datetime(2026,2,15), "MAINTENANCE FEE - CBC PARKING FACILITIES", "CORP-003", 35000, 0, "CBC Facilities", ""),
]
corp_balance = 2000000.00
for i, (dt, narr, ref, deb, cred, name, unit) in enumerate(corp_data):
    r = i + 2
    ws2.cell(r, 1, dt).number_format = date_fmt
    ws2.cell(r, 2, narr)
    ws2.cell(r, 3, ref)
    ws2.cell(r, 4, deb).number_format = num_fmt
    ws2.cell(r, 5, cred).number_format = num_fmt
    ws2.cell(r, 6, corp_balance - deb + cred if i == 0 else f"=F{r-1}-D{r}+E{r}")
    ws2.cell(r, 6).number_format = num_fmt
    ws2.cell(r, 7, unit)
    ws2.cell(r, 8, name)
    style_row(ws2, r, 8)

wb.save(os.path.join(OUT_DIR, "CBC_Master_Sheet.xlsx"))
print(f"Created: CBC_Master_Sheet.xlsx")

# ════════════════════════════════════════════════════════════════
# FILE 2: CBC_Escrow_Statement_Mar2026.xlsx
#
# REALISTIC bank statement — contains BOTH:
#   A) Old transactions that ARE already in the master sheet (agent should SKIP)
#   B) New transactions that are NOT in the master sheet (agent should MATCH & ADD)
#
# The agent must compare references/dates/amounts to determine which are new.
# ════════════════════════════════════════════════════════════════
wb2 = Workbook()
ws = wb2.active
ws.title = "Statement"

stmt_headers = ["Transaction Date", "Value Date", "Description", "Reference No",
                "Debit", "Credit", "Balance"]
for c, h in enumerate(stmt_headers, 1):
    ws.cell(1, c, h)
style_header(ws, 7)
for col, w in {'A':14,'B':14,'C':75,'D':25,'E':14,'F':14,'G':16}.items():
    ws.column_dimensions[col].width = w

# Mix of OLD (already in master) and NEW (not in master) transactions
# The bank statement covers Jan 5 → Mar 26, 2026
all_stmt_txns = [
    # ═══ OLD TRANSACTIONS (already in master sheet — agent should SKIP these) ═══
    (datetime(2026,1,5), datetime(2026,1,5),
     "OUTWARD CLEARINGUNIT 101 DEFAULT - ~026ABC001", "REF-H001", 0, 450000),
    (datetime(2026,1,8), datetime(2026,1,8),
     "INWARD REMITTANCETT REF: 530P4F86D98E100 AED 380000 FATIMA ABDULRAHMAN SAEED PO BOX 5432 DUBAI",
     "REF-H002", 0, 380000),
    (datetime(2026,1,12), datetime(2026,1,12),
     "OUTWARD CLEARINGUNIT 203 DEFAULT - ~026ABC003", "REF-H003", 0, 520000),
    (datetime(2026,1,15), datetime(2026,1,15),
     "INWARD REMITTANCETT REF: 044ABLC7890100 AED 490000 SARAH JOHNSON 45 PARK LANE LONDON",
     "REF-H004", 0, 490000),
    (datetime(2026,1,20), datetime(2026,1,20),
     "OUTWARD CLEARINGUNIT 305 DEFAULT - ~026ABC005", "REF-H005", 0, 410000),
    (datetime(2026,1,25), datetime(2026,1,25),
     "INWARD REMITTANCETT REF: 20260125001000 AED 375000 NADIA HASSAN IBRAHIM",
     "REF-H006", 0, 375000),
    (datetime(2026,2,1), datetime(2026,2,1),
     "OUTWARD CLEARINGUNIT 407 DEFAULT - ~026ABC007", "REF-H007", 0, 560000),
    (datetime(2026,2,5), datetime(2026,2,5),
     "INWARD REMITTANCETT REF: 530P5EB4FCD500 AED 430000 ELENA PETROVA MOSCOW RUSSIA",
     "REF-H008", 0, 430000),
    (datetime(2026,2,10), datetime(2026,2,10),
     "OUTWARD CLEARINGUNIT 509 DEFAULT - ~026ABC009", "REF-H009", 0, 510000),
    (datetime(2026,2,15), datetime(2026,2,15),
     "INWARD REMITTANCETT REF: 20260215005000 AED 395000 JAMES ROBERT WILSON FLAT 12 CHICAGO",
     "REF-H010", 0, 395000),

    # ═══ Also some non-unit debits the agent should ignore ═══
    (datetime(2026,2,20), datetime(2026,2,20),
     "BANK CHARGES - MONTHLY FEE", "REF-BANK01", 1500, 0),
    (datetime(2026,3,1), datetime(2026,3,1),
     "SWIFT CHARGES - INWARD REMITTANCE", "REF-BANK02", 250, 0),
    (datetime(2026,3,10), datetime(2026,3,10),
     "BANK CHARGES - STATEMENT FEE", "REF-BANK03", 100, 0),

    # ═══ NEW TRANSACTIONS (NOT in master — agent must detect, match, and add) ═══

    # 1. Easy: full name match → Unit 101 Q1 installment
    (datetime(2026,3,14), datetime(2026,3,14),
     "INWARD REMITTANCETT REF: 530P4F86D98E301 AED 56250 AHMED HASSAN AL MAKTOUM PO BOX 12345 DUBAI UAE",
     "STMT-301", 0, 56250),

    # 2. Easy: unit number in narration → Unit 102 Q1
    (datetime(2026,3,15), datetime(2026,3,15),
     "OUTWARD CLEARINGUNIT 102 DEFAULT - ~026STU901",
     "STMT-302", 0, 47500),

    # 3. Medium: "RASHIDI" missing "AL" prefix → Unit 203 Q1
    (datetime(2026,3,16), datetime(2026,3,16),
     "INWARD REMITTANCETT REF: 033DBLC2608235 AED 65000 MOHAMED KHALID RASHIDI CIT YWALK BUILDING DUBAI",
     "STMT-303", 0, 65000),

    # 4. Hard: IPP credit no name → Unit 204 Q1 (needs AI)
    (datetime(2026,3,17), datetime(2026,3,17),
     "IPP Customer CreditIPP 20260317ADC6B981204",
     "STMT-304", 0, 61250),

    # 5. Medium: "KH" abbreviation for KHALID → Unit 305 Q1
    (datetime(2026,3,18), datetime(2026,3,18),
     "INWARD REMITTANCETT REF: 20260318004944 AED 51250 KH OMAR AL MANSOURI",
     "STMT-305", 0, 51250),

    # 6. Easy: unit in narration → Unit 306 Q1
    (datetime(2026,3,19), datetime(2026,3,19),
     "OUTWARD CLEARINGUNIT 306 DEFAULT - ~026VWX234",
     "STMT-306", 0, 46875),

    # 7. Medium: "ALKETBI" no space → Unit 407 Q1
    (datetime(2026,3,20), datetime(2026,3,20),
     "INWARD REMITTANCETT REF: 530P5EB4FCD5BB AED 70000 RASHID SAEED ALKETBI",
     "STMT-307", 0, 70000),

    # 8. Hard: "PETROVNA" vs "PETROVA" → Unit 408 Q1
    (datetime(2026,3,21), datetime(2026,3,21),
     "INWARD REMITTANCETT REF: 044ABLC1234567 AED 53750 ELENA PETROVNA MOSCOW RUSSIA",
     "STMT-308", 0, 53750),

    # 9. Hard: "A M" initials → Unit 509 Q1
    (datetime(2026,3,22), datetime(2026,3,22),
     "INWARD REMITTANCETT REF: 20260322123456 AED 63750 A M AL SUWAIDI",
     "STMT-309", 0, 63750),

    # 10. Easy: unit in narration → Unit 510 Q1
    (datetime(2026,3,22), datetime(2026,3,22),
     "OUTWARD CLEARINGUNIT 510 DEFAULT - ~026YZA567",
     "STMT-310", 0, 49375),

    # 11. Hard: partial name "AHMED H MAKTOUM" → Unit 101 Q2
    (datetime(2026,3,23), datetime(2026,3,23),
     "INWARD REMITTANCETT REF: 789ABC123DEF01 AED 56250 AHMED H MAKTOUM",
     "STMT-311", 0, 56250),

    # 12. Medium: "FATIMA SAEED" dropped middle → Unit 102 Q2
    (datetime(2026,3,24), datetime(2026,3,24),
     "INWARD REMITTANCETT REF: 456DEF789GHI02 AED 47500 FATIMA SAEED",
     "STMT-312", 0, 47500),

    # 13. Hard: MOHAMMED vs MOHAMED + initial → Unit 203 Q2
    (datetime(2026,3,24), datetime(2026,3,24),
     "BANKNET AE420350000000123456789 12345678 MOHAMMED K AL RASHIDI",
     "STMT-313", 0, 65000),

    # 14. Medium: "NADIA IBRAHIM" dropped HASSAN → Unit 306 Q2
    (datetime(2026,3,25), datetime(2026,3,25),
     "INWARD REMITTANCETT REF: 111222333444 AED 46875 NADIA IBRAHIM",
     "STMT-314", 0, 46875),

    # 15. Hardest: blank narration → needs AI
    (datetime(2026,3,26), datetime(2026,3,26),
     "-",
     "STMT-315", 0, 51250),
]

balance = 5000000.00  # Same starting balance as master sheet
for i, (dt, vd, desc, ref, deb, cred) in enumerate(all_stmt_txns):
    r = i + 2
    ws.cell(r, 1, dt).number_format = date_fmt
    ws.cell(r, 2, vd).number_format = date_fmt
    ws.cell(r, 3, desc)
    ws.cell(r, 4, ref)
    ws.cell(r, 5, deb).number_format = num_fmt
    ws.cell(r, 6, cred).number_format = num_fmt
    balance = balance - deb + cred
    ws.cell(r, 7, balance).number_format = num_fmt
    style_row(ws, r, 7)

wb2.save(os.path.join(OUT_DIR, "CBC_Escrow_Statement_Mar2026.xlsx"))
print(f"Created: CBC_Escrow_Statement_Mar2026.xlsx")

print("\n=== Test Data Summary ===")
print(f"Master Sheet: 10 units, 10 BOOKING transactions")
print(f"Bank Statement: 28 total transactions")
print(f"  → 10 OLD (match master refs REF-H001..H010 — agent should SKIP)")
print(f"  → 3  BANK CHARGES (debits — agent should SKIP)")
print(f"  → 15 NEW Q1/Q2 installments (STMT-301..315 — agent should MATCH & ADD)")
print(f"\nExpected result after reconciliation:")
print(f"  → Master sheet grows from 10 to 25 rows (10 old + 15 new)")
print(f"  → 15 new rows have: unit number, account name, running balance formula")
print(f"  → Receipt column blank until SF receipt is generated → then 'Done'")
