from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
data_font = Font(name="Arial", size=10)
green_fill = PatternFill("solid", fgColor="C6EFCE")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
num_fmt = '#,##0.00'
date_fmt = 'DD/MM/YYYY'

def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font, cell.border = data_font, thin_border
        if fill: cell.fill = fill

# ================================================================
# MASTER SHEET - 10 units, booking payments only
# ================================================================
wb = Workbook()
ws1 = wb.active
ws1.title = "Updated Sheet_Escrow Account"
headers = ["Date", "Value Date", "Narration", "Reference", "Debit", "Credit",
           "Running Balance", "Unit No. / Remarks", "Account Name", "Receipt", "Status"]
for c, h in enumerate(headers, 1): ws1.cell(1, c, h)
style_header(ws1, 11)
for col, w in {'A':12,'B':12,'C':55,'D':22,'E':14,'F':14,'G':16,'H':18,'I':32,'J':10,'K':10}.items():
    ws1.column_dimensions[col].width = w

historical = [
    (datetime(2026,1,5), "OUTWARD CLEARINGUNIT 101 DEFAULT - ~026H001", "REF-H001", 0, 450000, "101", "Ahmed Hassan Al Maktoum"),
    (datetime(2026,1,10), "INWARD REMITTANCETT REF: 530PABC100 AED 380000 FATIMA ABDULRAHMAN SAEED PO BOX 5432 DUBAI", "REF-H002", 0, 380000, "102", "Fatima Abdulrahman Saeed"),
    (datetime(2026,1,15), "OUTWARD CLEARINGUNIT 203 DEFAULT - ~026H003", "REF-H003", 0, 520000, "203", "Mohamed Khalid Al Rashidi"),
    (datetime(2026,1,20), "INWARD REMITTANCETT REF: 044ABLC100 AED 490000 SARAH JOHNSON 45 PARK LANE LONDON", "REF-H004", 0, 490000, "204", "Sarah Johnson"),
    (datetime(2026,1,25), "OUTWARD CLEARINGUNIT 305 DEFAULT - ~026H005", "REF-H005", 0, 410000, "305", "Khalid Omar Al Mansouri"),
    (datetime(2026,2,1), "INWARD REMITTANCETT REF: 20260201001 AED 375000 NADIA HASSAN IBRAHIM", "REF-H006", 0, 375000, "306", "Nadia Hassan Ibrahim"),
    (datetime(2026,2,5), "OUTWARD CLEARINGUNIT 407 DEFAULT - ~026H007", "REF-H007", 0, 560000, "407", "Rashid Saeed Al Ketbi"),
    (datetime(2026,2,10), "INWARD REMITTANCETT REF: 530PDEF100 AED 430000 ELENA PETROVA MOSCOW RUSSIA", "REF-H008", 0, 430000, "408", "Elena Petrova"),
    (datetime(2026,2,15), "OUTWARD CLEARINGUNIT 509 DEFAULT - ~026H009", "REF-H009", 0, 510000, "509", "Ali Mohammed Al Suwaidi"),
    (datetime(2026,2,20), "INWARD REMITTANCETT REF: 20260220005 AED 395000 JAMES ROBERT WILSON FLAT 12 CHICAGO", "REF-H010", 0, 395000, "510", "James Robert Wilson"),
]

balance = 5000000.0
for i, (dt, narr, ref, deb, cred, unit, name) in enumerate(historical):
    r = i + 2
    ws1.cell(r, 1, dt).number_format = date_fmt
    ws1.cell(r, 2, dt).number_format = date_fmt
    ws1.cell(r, 3, narr)
    ws1.cell(r, 4, ref)
    ws1.cell(r, 5, deb).number_format = num_fmt
    ws1.cell(r, 6, cred).number_format = num_fmt
    if i == 0:
        ws1.cell(r, 7, balance - deb + cred).number_format = num_fmt
    else:
        ws1.cell(r, 7, f"=G{r-1}-E{r}+F{r}")
    ws1.cell(r, 7).number_format = num_fmt
    ws1.cell(r, 8, unit)
    ws1.cell(r, 9, name)
    ws1.cell(r, 10, "Done")
    ws1.cell(r, 11, "NEW")
    style_row(ws1, r, 11, green_fill)

# Corporate sheet (minimal)
ws2 = wb.create_sheet("Updated Sheet_Corporate")
corp_headers = ["Date", "Narration", "Reference", "Debit", "Credit", "Running Balance", "Unit No. / Remarks", "Account Name"]
for c, h in enumerate(corp_headers, 1): ws2.cell(1, c, h)
style_header(ws2, 8)

wb.save(os.path.join(OUT_DIR, "CBC_V2_Master.xlsx"))
print("Created: CBC_V2_Master.xlsx")

# ================================================================
# BANK STATEMENT - old + new transactions mixed
# Includes 5 transactions ONLY AI can match
# ================================================================
wb2 = Workbook()
ws = wb2.active
ws.title = "Statement"
stmt_headers = ["Transaction Date", "Value Date", "Description", "Reference No", "Debit", "Credit", "Balance"]
for c, h in enumerate(stmt_headers, 1): ws.cell(1, c, h)
style_header(ws, 7)
for col, w in {'A':14,'B':14,'C':75,'D':25,'E':14,'F':14,'G':16}.items():
    ws.column_dimensions[col].width = w

all_txns = [
    # === OLD (already in master - agent skips) ===
    (datetime(2026,1,5), "OUTWARD CLEARINGUNIT 101 DEFAULT - ~026H001", "REF-H001", 0, 450000),
    (datetime(2026,1,10), "INWARD REMITTANCETT REF: 530PABC100 AED 380000 FATIMA ABDULRAHMAN SAEED PO BOX 5432 DUBAI", "REF-H002", 0, 380000),
    (datetime(2026,1,15), "OUTWARD CLEARINGUNIT 203 DEFAULT - ~026H003", "REF-H003", 0, 520000),
    (datetime(2026,1,20), "INWARD REMITTANCETT REF: 044ABLC100 AED 490000 SARAH JOHNSON 45 PARK LANE LONDON", "REF-H004", 0, 490000),
    (datetime(2026,1,25), "OUTWARD CLEARINGUNIT 305 DEFAULT - ~026H005", "REF-H005", 0, 410000),
    (datetime(2026,2,1), "INWARD REMITTANCETT REF: 20260201001 AED 375000 NADIA HASSAN IBRAHIM", "REF-H006", 0, 375000),
    (datetime(2026,2,5), "OUTWARD CLEARINGUNIT 407 DEFAULT - ~026H007", "REF-H007", 0, 560000),
    (datetime(2026,2,10), "INWARD REMITTANCETT REF: 530PDEF100 AED 430000 ELENA PETROVA MOSCOW RUSSIA", "REF-H008", 0, 430000),
    (datetime(2026,2,15), "OUTWARD CLEARINGUNIT 509 DEFAULT - ~026H009", "REF-H009", 0, 510000),
    (datetime(2026,2,20), "INWARD REMITTANCETT REF: 20260220005 AED 395000 JAMES ROBERT WILSON FLAT 12 CHICAGO", "REF-H010", 0, 395000),
    # === BANK CHARGES (skip) ===
    (datetime(2026,2,28), "BANK CHARGES - MONTHLY ACCOUNT FEE", "BANK-001", 1500, 0),
    (datetime(2026,3,5), "SWIFT CHARGES - INWARD TT", "BANK-002", 250, 0),

    # === NEW TRANSACTIONS (agent must detect, match, add) ===

    # 1. EASY: unit in narration -> 102
    (datetime(2026,3,14), "OUTWARD CLEARINGUNIT 102 DEFAULT - ~026NEW01", "NEW-001", 0, 47500),
    # 2. EASY: full name -> 101
    (datetime(2026,3,14), "INWARD REMITTANCETT REF: 530PNEW002 AED 56250 AHMED HASSAN AL MAKTOUM PO BOX 12345 DUBAI UAE", "NEW-002", 0, 56250),
    # 3. MEDIUM: surname variant RASHIDI vs AL RASHIDI -> 203
    (datetime(2026,3,15), "INWARD REMITTANCETT REF: 033DBLNEW03 AED 65000 MOHAMED KHALID RASHIDI CIT BUILDING DUBAI", "NEW-003", 0, 65000),
    # 4. EASY: unit in narration -> 306
    (datetime(2026,3,16), "OUTWARD CLEARINGUNIT 306 DEFAULT - ~026NEW04", "NEW-004", 0, 46875),
    # 5. MEDIUM: compound surname ALKETBI vs AL KETBI -> 407
    (datetime(2026,3,17), "INWARD REMITTANCETT REF: 530PNEW005 AED 70000 RASHID SAEED ALKETBI", "NEW-005", 0, 70000),
    # 6. EASY: unit in narration -> 510
    (datetime(2026,3,18), "OUTWARD CLEARINGUNIT 510 DEFAULT - ~026NEW06", "NEW-006", 0, 49375),
    # 7. MEDIUM: partial name FATIMA SAEED (dropped ABDULRAHMAN) -> 102
    (datetime(2026,3,19), "INWARD REMITTANCETT REF: 456DEFNEW07 AED 47500 FATIMA SAEED", "NEW-007", 0, 47500),
    # 8. MEDIUM: MOHAMMED vs MOHAMED variant + initial K -> 203
    (datetime(2026,3,20), "BANKNET AE420350000000987654321 99887766 MOHAMMED K AL RASHIDI", "NEW-008", 0, 65000),
    # 9. MEDIUM: dropped middle name NADIA IBRAHIM -> 306
    (datetime(2026,3,21), "INWARD REMITTANCETT REF: 111NEW009 AED 46875 NADIA IBRAHIM", "NEW-009", 0, 46875),
    # 10. MEDIUM: abbreviated KH OMAR -> 305
    (datetime(2026,3,22), "INWARD REMITTANCETT REF: 20260322NEW10 AED 51250 KH OMAR AL MANSOURI", "NEW-010", 0, 51250),

    # === HARD: These need AI matching ===

    # 11. HARD: IPP with unit hint 204 in reference -> 204
    (datetime(2026,3,23), "IPP Customer CreditIPP 20260323ADC6B981204", "NEW-011", 0, 61250),
    # 12. HARD: Russian patronymic PETROVNA vs PETROVA -> 408
    (datetime(2026,3,23), "INWARD REMITTANCETT REF: 044ABLNEW12 AED 53750 ELENA PETROVNA MOSCOW RUSSIA", "NEW-012", 0, 53750),
    # 13. HARD: Only initials A M + surname -> 509
    (datetime(2026,3,24), "INWARD REMITTANCETT REF: 20260324NEW13 AED 63750 A M AL SUWAIDI", "NEW-013", 0, 63750),
    # 14. HARD: Company name paying on behalf -> 204 (AI should deduce from amount)
    (datetime(2026,3,25), "INWARD REMITTANCETT REF: CORP789NEW14 AED 61250 PARK LANE INVESTMENTS LLC LONDON", "NEW-014", 0, 61250),
    # 15. HARDEST: blank narration, only amount -> AI must use amount pattern
    (datetime(2026,3,26), "-", "NEW-015", 0, 51250),
]

balance = 5000000.0
for i, (dt, desc, ref, deb, cred) in enumerate(all_txns):
    r = i + 2
    ws.cell(r, 1, dt).number_format = date_fmt
    ws.cell(r, 2, dt).number_format = date_fmt
    ws.cell(r, 3, desc)
    ws.cell(r, 4, ref)
    ws.cell(r, 5, deb).number_format = num_fmt
    ws.cell(r, 6, cred).number_format = num_fmt
    balance = balance - deb + cred
    ws.cell(r, 7, balance).number_format = num_fmt
    style_row(ws, r, 7)

wb2.save(os.path.join(OUT_DIR, "CBC_V2_Escrow.xlsx"))
print("Created: CBC_V2_Escrow.xlsx")

print("\n=== V2 Test Data ===")
print("Master: 10 units, 10 booking payments")
print("Statement: 27 transactions total")
print("  10 old (skip) + 2 bank charges (skip) + 15 new")
print("\nNew transaction breakdown:")
print("  EASY (unit in narration):  3  (#1, #4, #6)")
print("  EASY (full name):          1  (#2)")
print("  MEDIUM (name variants):    5  (#3, #5, #7, #8, #9, #10)")
print("  HARD (needs rule+AI):      5  (#11 IPP, #12 PETROVNA, #13 initials, #14 company, #15 blank)")
print("\nExpected: 10 rule-based + 3 rule-based-hard + 2 AI-only = 15/15")
