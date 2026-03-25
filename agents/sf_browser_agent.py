"""
Salesforce Browser Automation Agent — Uses Chrome browser session to perform
actions in Salesforce without API credentials.

Requires: Claude in Chrome MCP tools (runs from Claude Code/Cowork session)

TWO SALESFORCE APPS (same org: momentum-ability-3447):
  - "Fam Properties" (fam app): Used for updating payment amounts only
  - "Fam Properties (Revamp)": Used for Generate Invoice, Generate Statement,
    Send Email, and all receipt/SOA operations

Workflow per matched unit:
1. Switch to "fam" app → Search unit → Payments tab → Update Amount Paid
2. Switch to "fam revamp" app → Same unit → Payments tab → Find payment record
3. Click "Generate Invoice" to create receipt
4. Go to Account Statements tab → verify/generate statement
5. Send email to buyer via SF Activity panel (with receipt + statement attached)
6. Mark receipt as "Done" in master sheet
"""

import time
import json
from dataclasses import dataclass, field, asdict
from typing import Optional


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class SFAction:
    """A single Salesforce browser action."""
    unit_no: str = ""
    payment_id: str = ""        # e.g., BP-03041
    action_type: str = ""       # "generate_invoice", "update_payment", "send_email"
    status: str = "pending"     # "pending", "in_progress", "done", "error", "dry_run"
    message: str = ""
    details: dict = field(default_factory=dict)

    def to_dict(self):
        return asdict(self)


@dataclass
class SFBrowserConfig:
    """Configuration for the SF browser agent."""
    sf_base_url: str = "https://momentum-ability-3447.lightning.force.com"
    # Two SF apps in the same org:
    # "fam" app — for updating payment amounts
    fam_app_id: str = "06mNM000000gMB1"  # Fam Properties app ID (update if different)
    # "fam revamp" app — for Generate Invoice, Statement, Email
    revamp_app_id: str = "06mNM000000YtKv"  # Fam Properties (Revamp) app ID (update if different)
    wait_seconds: float = 3.0   # Wait time between page loads
    confirm_each: bool = False  # If True, pause and confirm before each action


class SalesforceBrowserAgent:
    """
    Automates Salesforce operations through the browser using Chrome MCP tools.

    This agent is designed to be called from a Claude Code session where
    the Chrome MCP tools are available. It does NOT run on Vercel.

    Usage from Claude Code:
        agent = SalesforceBrowserAgent(tab_id=12345)
        results = agent.process_units(needs_receipt_list)
    """

    def __init__(self, tab_id: int = None, config: SFBrowserConfig = None):
        self.tab_id = tab_id
        self.config = config or SFBrowserConfig()
        self.actions_log: list[SFAction] = []

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def process_units(self, units: list[dict]) -> list[SFAction]:
        """
        Process a list of units that need SF actions.
        Each unit dict should have: unit_no, credit, account_name, date, etc.

        This method returns the planned actions. Actual execution happens
        via the execute_* methods which use Chrome MCP tools.
        """
        self.actions_log = []

        for unit_info in units:
            unit_no = unit_info.get("unit_no", "")
            amount = unit_info.get("credit", 0)
            client = unit_info.get("account_name", "")

            if not unit_no:
                continue

            # Action 1: Update payment in fam app
            self.actions_log.append(SFAction(
                unit_no=unit_no,
                action_type="update_payment",
                status="pending",
                message=f"Update payment in fam app for Unit {unit_no} — AED {amount:,.2f}",
                details={"amount": amount, "client": client, "app": "fam"},
            ))

            # Action 2: Generate invoice in fam revamp app
            self.actions_log.append(SFAction(
                unit_no=unit_no,
                action_type="generate_invoice",
                status="pending",
                message=f"Generate invoice in Revamp app for Unit {unit_no}",
                details={"amount": amount, "client": client, "app": "fam_revamp"},
            ))

            # Action 3: Generate statement in fam revamp app
            self.actions_log.append(SFAction(
                unit_no=unit_no,
                action_type="generate_statement",
                status="pending",
                message=f"Generate statement in Revamp app for Unit {unit_no}",
                details={"app": "fam_revamp"},
            ))

            # Action 4: Send email from fam revamp app
            self.actions_log.append(SFAction(
                unit_no=unit_no,
                action_type="send_email",
                status="pending",
                message=f"Send receipt + statement email to buyer of Unit {unit_no}",
                details={"client": client, "attachments": ["receipt", "statement"], "app": "fam_revamp"},
            ))

        return self.actions_log

    # ------------------------------------------------------------------
    # Browser automation steps (use Chrome MCP tools)
    # These are designed to be called by Claude Code with MCP access
    # ------------------------------------------------------------------

    def get_search_url(self, unit_no: str) -> str:
        """Get the SF search URL for a unit number."""
        return f"{self.config.sf_base_url}/lightning/o/Inventory_Unit__c/list?filterName=Recent"

    def get_unit_url(self, unit_no: str, project: str = "CENTURY") -> str:
        """
        Construct the SF search URL to find a unit.
        Note: The actual SF record ID is needed — we search by unit name.
        """
        # SF global search URL
        return f"{self.config.sf_base_url}/lightning/o/Inventory_Unit__c/list?filterName=Recent&q={unit_no}"

    # ------------------------------------------------------------------
    # Step-by-step browser instructions (for Claude to execute)
    # ------------------------------------------------------------------

    def get_instructions_for_unit(self, unit_no: str, amount: float = 0, client: str = "") -> list[dict]:
        """
        Return step-by-step browser instructions for processing a unit.
        Uses TWO Salesforce apps:
        - "fam" app: Update payment amount
        - "fam revamp" app: Generate Invoice, Statement, Send Email
        """
        base = self.config.sf_base_url
        steps = []

        # ============================================================
        # PHASE 1: Update payment in "fam" app
        # ============================================================

        steps.append({
            "step": 1,
            "phase": "fam_app",
            "action": "switch_app",
            "description": "Switch to 'Fam Properties' app (fam app)",
            "hint": "Click the App Launcher (9-dot grid icon top-left) → search 'Fam Properties' → click it (NOT 'Fam Properties (Revamp)')",
            "url": f"{base}/lightning/app/standard__LightningSalesConsole",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 2,
            "phase": "fam_app",
            "action": "search_unit",
            "description": f"Search for Unit {unit_no} in the fam app",
            "hint": f"Use the global search bar at the top → type '{unit_no}' → look for Inventory record → click it",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 3,
            "phase": "fam_app",
            "action": "go_to_payments",
            "description": "Click the 'Payments' tab on the unit record",
            "hint": "Tab labeled 'Payments' between 'Details' and 'Account Statements'",
            "wait_after": self.config.wait_seconds + 1,
        })

        steps.append({
            "step": 4,
            "phase": "fam_app",
            "action": "find_payment",
            "description": f"Find the payment record for AED {amount:,.2f} (or the next unpaid/overdue one)",
            "hint": "Look at Sub Total, Status, and Payment Detail columns. Click the BP-XXXXX link",
            "amount": amount,
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 5,
            "phase": "fam_app",
            "action": "update_payment",
            "description": f"Click 'Edit' → Update 'Amount Paid' field to include AED {amount:,.2f} → Click 'Save'",
            "hint": "The Amount Paid field is in the Payment Information section. Add the new amount to the existing Amount Paid value.",
            "amount": amount,
            "wait_after": self.config.wait_seconds,
            "confirm": True,
        })

        # ============================================================
        # PHASE 2: Generate Receipt + Statement + Email in "fam revamp" app
        # ============================================================

        steps.append({
            "step": 6,
            "phase": "fam_revamp",
            "action": "switch_app",
            "description": "Switch to 'Fam Properties (Revamp)' app",
            "hint": "Click the App Launcher (9-dot grid icon) → search 'Fam Properties' → click 'Fam Properties (Rev...)' (the Revamp version)",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 7,
            "phase": "fam_revamp",
            "action": "search_unit",
            "description": f"Search for Unit {unit_no} in the Revamp app",
            "hint": f"Global search → type '{unit_no}' → click the Inventory record",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 8,
            "phase": "fam_revamp",
            "action": "go_to_payments",
            "description": "Click 'Payments' tab → find the same payment record",
            "hint": "Same BP-XXXXX record from Phase 1",
            "wait_after": self.config.wait_seconds + 1,
        })

        steps.append({
            "step": 9,
            "phase": "fam_revamp",
            "action": "generate_invoice",
            "description": "Click 'Generate Invoice' button to create the receipt",
            "hint": "Button in the top-right area of the payment record, next to 'Edit' and 'Waived Off Compensation'",
            "wait_after": self.config.wait_seconds + 2,
            "confirm": True,
        })

        steps.append({
            "step": 10,
            "phase": "fam_revamp",
            "action": "confirm_dialog",
            "description": "Handle any confirmation dialog — click OK/Confirm/Save",
            "hint": "Look for a modal popup with a confirm button",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 11,
            "phase": "fam_revamp",
            "action": "go_back_to_unit",
            "description": f"Click the Unit '{unit_no}' link to go back to the inventory record",
            "hint": "Click the Inventory field value (the unit number link)",
            "wait_after": self.config.wait_seconds,
        })

        steps.append({
            "step": 12,
            "phase": "fam_revamp",
            "action": "go_to_statements",
            "description": "Click 'Account Statements' tab to verify/generate statement",
            "hint": "Tab between 'Payments' and 'Documents'",
            "wait_after": self.config.wait_seconds + 1,
        })

        steps.append({
            "step": 13,
            "phase": "fam_revamp",
            "action": "send_email",
            "description": f"Send email to buyer ({client}) with receipt + statement attached",
            "hint": "Click the email icon (envelope) in the Activity panel on the right side of the page",
            "email_details": {
                "to": client or f"Buyer of Unit {unit_no}",
                "subject": f"[fam Properties] Payment Receipt & Statement — Century Unit No.{unit_no}",
                "body": (
                    f"Dear Valued Client,\n\n"
                    f"Good day!\n\n"
                    f"Please find attached the payment receipt for the amount paid towards "
                    f"Century Unit No.{unit_no}. Also attached is the updated Statement of Account "
                    f"for your reference.\n\n"
                    f"Kindly acknowledge receipt of this email.\n\n"
                    f"Regards"
                ),
                "attachments": ["Payment Receipt PDF", "Statement of Account PDF"],
            },
            "wait_after": self.config.wait_seconds,
            "confirm": True,
        })

        return steps

    # ------------------------------------------------------------------
    # Batch processing plan
    # ------------------------------------------------------------------

    def generate_batch_plan(self, needs_receipt: list[dict]) -> dict:
        """
        Generate a complete execution plan for all units needing receipts.
        Returns a structured plan that can be reviewed before execution.
        """
        plan = {
            "total_units": len(needs_receipt),
            "dry_run": self.config.dry_run,
            "sf_base_url": self.config.sf_base_url,
            "units": [],
        }

        for item in needs_receipt:
            unit_no = item.get("unit_no", "")
            if not unit_no:
                continue

            unit_plan = {
                "unit_no": unit_no,
                "amount": item.get("credit", 0),
                "client": item.get("account_name", ""),
                "date": item.get("date", ""),
                "row_in_master": item.get("row", ""),
                "sheet": item.get("sheet", ""),
                "steps": self.get_instructions_for_unit(
                    unit_no=unit_no,
                    amount=item.get("credit", 0),
                    client=item.get("account_name", ""),
                ),
                "actions": [
                    {
                        "type": "update_payment",
                        "app": "fam",
                        "description": "Update Amount Paid in 'Fam Properties' app",
                    },
                    {
                        "type": "generate_invoice",
                        "app": "fam_revamp",
                        "description": "Generate Invoice/Receipt in 'Fam Properties (Revamp)' app",
                    },
                    {
                        "type": "generate_statement",
                        "app": "fam_revamp",
                        "description": "Verify/Generate Statement of Account in Revamp app",
                    },
                    {
                        "type": "send_email",
                        "app": "fam_revamp",
                        "description": "Send receipt + statement email to buyer from Revamp app",
                    },
                    {
                        "type": "mark_master_sheet",
                        "app": "local",
                        "description": f"Mark receipt column as 'Done' in master sheet row {item.get('row', '?')}",
                    },
                ],
            }
            plan["units"].append(unit_plan)

        return plan

    # ------------------------------------------------------------------
    # Master sheet receipt column update
    # ------------------------------------------------------------------

    @staticmethod
    def mark_receipts_done(master_path: str, output_path: str, completed_units: list[dict]) -> dict:
        """
        Update the receipt column in the master sheet for completed units.
        completed_units: list of {"row": int, "sheet": str, "unit_no": str}
        """
        import openpyxl
        from openpyxl.styles import PatternFill

        wb = openpyxl.load_workbook(master_path)
        done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        updated = 0

        for item in completed_units:
            sheet_name = item.get("sheet", "")
            row = item.get("row", 0)
            if not sheet_name or not row or sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Receipt column: Col J (10) for Escrow, doesn't exist for Corporate
            if "Escrow" in sheet_name:
                receipt_col = 10
            else:
                continue  # Corporate doesn't have a receipt column

            ws.cell(row, receipt_col, "Done")
            ws.cell(row, receipt_col).fill = done_fill
            updated += 1

        wb.save(output_path)
        return {"updated": updated, "output_path": output_path}

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------

    def get_summary(self) -> dict:
        """Get a summary of all planned/executed actions."""
        by_type = {}
        by_status = {}
        for a in self.actions_log:
            by_type[a.action_type] = by_type.get(a.action_type, 0) + 1
            by_status[a.status] = by_status.get(a.status, 0) + 1

        return {
            "total_actions": len(self.actions_log),
            "by_type": by_type,
            "by_status": by_status,
            "actions": [a.to_dict() for a in self.actions_log],
        }


# ---------------------------------------------------------------------------
# Convenience function for Claude Code execution
# ---------------------------------------------------------------------------

def plan_sf_automation(needs_receipt: list[dict]) -> dict:
    """
    Generate an SF browser automation plan from the needs_receipt list.
    Uses dual-app approach: fam app for payments, fam revamp for receipts/emails.

    Usage in Claude Code/Cowork:
        from agents.sf_browser_agent import plan_sf_automation
        plan = plan_sf_automation(needs_receipt_data)
        print(json.dumps(plan, indent=2))
    """
    config = SFBrowserConfig()
    agent = SalesforceBrowserAgent(config=config)
    return agent.generate_batch_plan(needs_receipt)


def mark_receipts_in_master(
    master_path: str, output_path: str, completed_units: list[dict]
) -> dict:
    """
    Mark receipt column as 'Done' for completed units.

    Usage in Claude Code:
        from agents.sf_browser_agent import mark_receipts_in_master
        result = mark_receipts_in_master('Century_Updated.xlsx', 'output.xlsx', units)
    """
    return SalesforceBrowserAgent.mark_receipts_done(
        master_path, output_path, completed_units
    )
