"""
Daily Reconciler -- The core engine that Amit runs daily.

Input:  Two new bank statements (Escrow + Corporate) from Geo
        + The master Excel sheet with historical reconciled data

Process:
1. Load master sheet to build knowledge base (unit->name, ref->unit, name->unit)
2. Parse new Escrow and Corporate statements
3. Find NEW transactions (not already in master sheet, matched by Transaction Reference)
4. Match new transactions to units using: knowledge base + narration patterns + cross-reference
5. Append matched new transactions to the master sheet
6. Generate receipt PDFs for each matched new credit

Output: Updated master Excel + receipt PDFs + summary
"""

import os
import re
import copy
import json
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field, asdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from processors.bank_statement_parser import BankStatementParser, Transaction
from processors.pdf_generator import generate_receipt
from config.payment_settings import NAME_VARIANTS

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False


@dataclass
class NewTransaction:
    """A new transaction found in the daily statement that's not in the master sheet."""
    account: str = ""           # "escrow" or "corporate"
    date: str = ""
    value_date: str = ""
    narration: str = ""
    reference: str = ""
    debit: float = 0.0
    credit: float = 0.0
    balance: float = 0.0
    # Matching results
    unit_no: str = ""
    account_name: str = ""      # Client/buyer name
    match_method: str = ""
    match_confidence: float = 0.0
    receipt_generated: bool = False
    receipt_path: str = ""
    alternative_matches: list = field(default_factory=list)  # top 3 alt matches: [{unit_no, account_name, confidence, method}]

    def to_dict(self):
        return asdict(self)


class DailyReconciler:
    """
    The daily reconciliation engine.
    Replaces Amit's manual process of checking bank statements and updating the master sheet.
    """

    def __init__(self):
        # Knowledge base built from master sheet
        self.kb_ref_to_unit: dict[str, str] = {}       # transaction_ref -> unit
        self.kb_ref_to_name: dict[str, str] = {}       # transaction_ref -> client name
        self.kb_name_to_unit: dict[str, list[str]] = {}   # normalized_name -> [unit, ...]
        self.kb_unit_to_name: dict[str, list[str]] = {}  # unit -> [client name, ...]
        self.existing_refs_escrow: set[str] = set()     # refs already in escrow sheet
        self.existing_refs_corporate: set[str] = set()  # refs already in corporate sheet
        self.latest_date_escrow: datetime = None        # latest date in escrow sheet
        self.latest_date_corporate: datetime = None     # latest date in corporate sheet

        self.new_transactions: list[NewTransaction] = []
        self.receipts_generated: list[dict] = []

    # =================================================================
    # STEP 1: Load master sheet and build knowledge base
    # =================================================================

    def load_master_sheet(self, master_path: str) -> dict:
        """Load the master Excel and build the knowledge base from existing data."""
        wb = openpyxl.load_workbook(master_path, data_only=True)

        stats = {"escrow_rows": 0, "corporate_rows": 0, "units_known": 0, "names_known": 0}

        # Load Escrow sheet
        if "Updated Sheet_Escrow Account" in wb.sheetnames:
            ws = wb["Updated Sheet_Escrow Account"]
            stats["escrow_rows"] = self._load_sheet_to_kb(ws, "escrow", has_account_name=True)

        # Load Corporate sheet
        if "Updated Sheet_Corporate" in wb.sheetnames:
            ws = wb["Updated Sheet_Corporate"]
            stats["corporate_rows"] = self._load_sheet_to_kb(ws, "corporate", has_account_name=True)

        # Also check Corporate Recon sheet
        if "Corporate Recon as of 250524" in wb.sheetnames:
            ws = wb["Corporate Recon as of 250524"]
            self._load_recon_sheet(ws)

        stats["units_known"] = len(self.kb_unit_to_name)
        stats["names_known"] = len(self.kb_name_to_unit)
        stats["refs_escrow"] = len(self.existing_refs_escrow)
        stats["refs_corporate"] = len(self.existing_refs_corporate)
        stats["cutoff_escrow"] = self.latest_date_escrow.strftime("%d-%m-%Y") if self.latest_date_escrow else "N/A"
        stats["cutoff_corporate"] = self.latest_date_corporate.strftime("%d-%m-%Y") if self.latest_date_corporate else "N/A"

        # Find unreconciled rows and try to match them
        self._unreconciled_rows = []
        self._needs_receipt = []  # Rows with unit but no receipt
        self._scan_unreconciled(wb)
        stats["unreconciled_found"] = len(self._unreconciled_rows)
        stats["unreconciled_matched"] = sum(1 for u in self._unreconciled_rows if u.get("unit_no"))
        stats["needs_receipt"] = len(self._needs_receipt)

        return stats

    def _scan_unreconciled(self, wb):
        """Scan master sheet for rows without a unit assignment and try to match them.
        Uses ALL matching methods: unit extraction, KB name, fuzzy, substring, narration scan."""
        from processors.bank_statement_parser import BankStatementParser
        parser = BankStatementParser()

        for sheet_name, unit_col, narr_col, credit_col, name_col, receipt_col, acct_type in [
            ("Updated Sheet_Escrow Account", 8, 3, 6, 9, 10, "escrow"),
            ("Updated Sheet_Corporate", 7, 2, 5, 8, None, "corporate"),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for r in range(2, ws.max_row + 1):
                credit = ws.cell(r, credit_col).value
                if not credit:
                    continue
                try:
                    credit_float = float(str(credit).replace(",", ""))
                except (ValueError, TypeError):
                    continue
                if credit_float <= 0:
                    continue

                # Date filter: only process 2026+ transactions
                date_val = ws.cell(r, 1).value
                parsed_date = self._parse_date(date_val)
                # Only process transactions from 22-Mar-2026 onwards
                cutoff = datetime(2026, 3, 22)
                if parsed_date and parsed_date < cutoff:
                    continue

                unit_val = ws.cell(r, unit_col).value
                unit_str = str(unit_val).strip() if unit_val else ""

                # Check receipt status
                receipt_val = ""
                if receipt_col:
                    receipt_val = str(ws.cell(r, receipt_col).value or "").strip()

                # Skip rows that already have a unit assigned
                has_unit = unit_str and unit_str not in ("0", "0.0", "None", "", "UNMATCHED")
                if has_unit:
                    # Track rows with unit but no receipt (for SF receipt generation)
                    if receipt_col and not receipt_val:
                        self._needs_receipt.append({
                            "row": r,
                            "sheet": sheet_name,
                            "account": acct_type,
                            "date": str(date_val) if date_val else "",
                            "credit": credit_float,
                            "unit_no": unit_str.split(".")[0].split("-")[0].split(" ")[0],
                            "account_name": str(ws.cell(r, name_col).value or "") if name_col else "",
                        })
                    continue

                narr = str(ws.cell(r, narr_col).value or "")

                # === Multi-method matching ===
                unit = ""
                conf = 0.0
                match_method = ""
                name = ""
                alt_matches = []

                # Method 1: Unit number in narration
                unit, conf = parser.extract_unit_from_description(narr)
                if unit and conf >= 0.80:
                    match_method = "unit_in_narration"
                else:
                    unit = ""
                    conf = 0.0

                # Method 2: Extract name + KB lookup
                if not unit:
                    name = parser.extract_name_from_description(narr)
                    if name and len(name) > 3:
                        normalized = self._normalize_name(name)
                        # Exact KB match
                        if normalized in self.kb_name_to_unit:
                            unit = self.kb_name_to_unit[normalized][0]
                            conf = 0.85
                            match_method = "name_from_kb"
                            alt_matches = self._collect_fuzzy_alternatives(normalized, unit)
                        else:
                            # FIX 1: Best-match substring (collect ALL, pick best overlap)
                            best_sub_unit = ""
                            best_sub_ratio = 0.0
                            for kb_name, kb_units in self.kb_name_to_unit.items():
                                if len(normalized) > 5 and len(kb_name) > 5:
                                    if normalized in kb_name or kb_name in normalized:
                                        overlap = min(len(normalized), len(kb_name))
                                        total = max(len(normalized), len(kb_name))
                                        ratio = overlap / total if total > 0 else 0
                                        if ratio > best_sub_ratio:
                                            best_sub_ratio = ratio
                                            best_sub_unit = kb_units[0]
                            if best_sub_unit:
                                unit = best_sub_unit
                                conf = 0.82
                                match_method = "name_substring"
                                alt_matches = self._collect_fuzzy_alternatives(normalized, unit)

                # Method 3: Fuzzy name match with Arabic variants
                if not unit and name and len(name) > 3:
                    normalized = self._normalize_name(name)
                    best_unit, best_score, all_candidates = self._fuzzy_name_match(normalized, return_all=True)
                    if best_unit and best_score >= 0.55:
                        unit = best_unit
                        conf = round(min(0.80, best_score * 0.85), 2)
                        match_method = "name_fuzzy"
                        # top 3 alternatives excluding chosen
                        seen = {best_unit}
                        for c in all_candidates:
                            if c["unit_no"] not in seen:
                                seen.add(c["unit_no"])
                                alt_matches.append(c)
                            if len(alt_matches) >= 3:
                                break
                    elif all_candidates:
                        # Unmatched but still provide alternatives
                        seen = set()
                        for c in all_candidates:
                            if c["unit_no"] not in seen:
                                seen.add(c["unit_no"])
                                alt_matches.append(c)
                            if len(alt_matches) >= 3:
                                break

                # Method 4: Direct narration scan against KB names (FIX 2: scored approach)
                if not unit:
                    narr_upper = narr.upper()
                    best_narr_unit = ""
                    best_narr_score = 0.0
                    best_narr_kb_name = ""
                    for kb_name, kb_units in self.kb_name_to_unit.items():
                        parts = [p for p in kb_name.split() if len(p) > 2]
                        if len(parts) < 2:
                            continue
                        last = parts[-1]
                        if last in narr_upper and any(p in narr_upper for p in parts[:-1]):
                            matching_parts = sum(1 for p in parts if p in narr_upper)
                            score = matching_parts / len(parts)
                            if score > best_narr_score:
                                best_narr_score = score
                                best_narr_unit = kb_units[0]
                                best_narr_kb_name = kb_name
                    if best_narr_unit:
                        unit = best_narr_unit
                        conf = 0.75
                        match_method = "name_in_narration"
                        names_list = self.kb_unit_to_name.get(best_narr_unit, [])
                        name = names_list[0] if names_list else ""
                        if name and len(name) > 3:
                            alt_matches = self._collect_fuzzy_alternatives(
                                self._normalize_name(name), unit)

                if not name and unit:
                    names_list = self.kb_unit_to_name.get(unit, [])
                    name = names_list[0] if names_list else ""

                self._unreconciled_rows.append({
                    "row": r,
                    "sheet": sheet_name,
                    "account": acct_type,
                    "date": str(date_val) if date_val else "",
                    "narration": narr[:200],
                    "credit": credit_float,
                    "unit_no": unit,
                    "account_name": name or "",
                    "match_method": match_method,
                    "match_confidence": conf if unit else 0.0,
                    "receipt_status": receipt_val,
                    "alternative_matches": alt_matches,
                })

    def _load_sheet_to_kb(self, ws, account_type: str, has_account_name: bool = False) -> int:
        """Load a sheet and extract knowledge base entries."""
        rows_loaded = 0

        for r in range(2, ws.max_row + 1):
            # Get transaction reference (column D for escrow, column C for corporate)
            if account_type == "escrow":
                date_val = ws.cell(r, 1).value  # Column A
                ref_val = ws.cell(r, 4).value   # Column D
                unit_val = ws.cell(r, 8).value  # Column H: "Unit No. / Remarks"
                name_val = ws.cell(r, 9).value  # Column I: "Account Name"
            else:  # corporate
                date_val = ws.cell(r, 1).value  # Column A
                ref_val = ws.cell(r, 3).value   # Column C
                unit_val = ws.cell(r, 7).value  # Column G: "Unit No. / Remarks"
                name_val = ws.cell(r, 8).value  # Column H: "Account Name"

            if not ref_val:
                continue

            ref = str(ref_val).strip()
            if not ref:
                continue

            # Track latest date
            parsed_date = self._parse_date(date_val)
            if parsed_date:
                if account_type == "escrow":
                    if not self.latest_date_escrow or parsed_date > self.latest_date_escrow:
                        self.latest_date_escrow = parsed_date
                else:
                    if not self.latest_date_corporate or parsed_date > self.latest_date_corporate:
                        self.latest_date_corporate = parsed_date

            # Track existing refs
            if account_type == "escrow":
                self.existing_refs_escrow.add(ref)
            else:
                self.existing_refs_corporate.add(ref)

            rows_loaded += 1

            # Extract unit number
            unit = ""
            if unit_val:
                unit_str = str(unit_val).strip()
                # Handle formats: "1701.0", "802 - Johanna", "1308 - Sagal", "1002.0"
                m = re.match(r'(\d{3,4})', unit_str)
                if m:
                    unit = m.group(1)

            # Extract client name
            name = ""
            if name_val:
                name = str(name_val).strip().split('\n')[0]  # Take first line

            # Build knowledge base
            if unit:
                self.kb_ref_to_unit[ref] = unit
                if name:
                    self.kb_ref_to_name[ref] = name
                    # Append to list (don't overwrite)
                    if unit not in self.kb_unit_to_name:
                        self.kb_unit_to_name[unit] = []
                    if name not in self.kb_unit_to_name[unit]:
                        self.kb_unit_to_name[unit].append(name)
                    normalized = self._normalize_name(name)
                    if len(normalized) > 3:
                        if normalized not in self.kb_name_to_unit:
                            self.kb_name_to_unit[normalized] = []
                        if unit not in self.kb_name_to_unit[normalized]:
                            self.kb_name_to_unit[normalized].append(unit)

        return rows_loaded

    def _load_recon_sheet(self, ws):
        """Load Corporate Recon sheet for additional unit-name mappings."""
        for r in range(2, ws.max_row + 1):
            remarks = ws.cell(r, 8).value  # Column H: Remarks "Unit No. 1308 - Sivan Hanouchian"
            ref = ws.cell(r, 4).value       # Column D: Transaction Reference

            if not remarks:
                continue

            remarks_str = str(remarks).strip()
            m = re.match(r'Unit\s*No\.?\s*(\d{3,4})\s*-\s*(.+)', remarks_str)
            if m:
                unit = m.group(1)
                name = m.group(2).strip()
                if unit and name:
                    if unit not in self.kb_unit_to_name:
                        self.kb_unit_to_name[unit] = []
                    if name not in self.kb_unit_to_name[unit]:
                        self.kb_unit_to_name[unit].append(name)
                    normalized = self._normalize_name(name)
                    if len(normalized) > 3:
                        if normalized not in self.kb_name_to_unit:
                            self.kb_name_to_unit[normalized] = []
                        if unit not in self.kb_name_to_unit[normalized]:
                            self.kb_name_to_unit[normalized].append(unit)
                    if ref:
                        self.kb_ref_to_unit[str(ref).strip()] = unit
                        self.kb_ref_to_name[str(ref).strip()] = name

    # =================================================================
    # STEP 2: Parse new statements and find new transactions
    # =================================================================

    def process_new_statements(
        self,
        escrow_path: str = None,
        corporate_path: str = None,
    ) -> dict:
        """Parse new bank statements and find transactions not in the master sheet."""
        results = {"escrow_new": 0, "corporate_new": 0, "total_new": 0}

        if escrow_path:
            parser = BankStatementParser()
            result = parser.parse(escrow_path)
            if not result.get("error"):
                new_count = self._find_new_transactions(parser, "escrow")
                results["escrow_new"] = new_count
                results["escrow_total"] = result["total_transactions"]

        if corporate_path:
            parser = BankStatementParser()
            result = parser.parse(corporate_path)
            if not result.get("error"):
                new_count = self._find_new_transactions(parser, "corporate")
                results["corporate_new"] = new_count
                results["corporate_total"] = result["total_transactions"]

        results["total_new"] = len(self.new_transactions)
        return results

    def _find_new_transactions(self, parser: BankStatementParser, account_type: str) -> int:
        """Find transactions not already in the master sheet. Only looks at transactions AFTER the latest date."""
        existing_refs = self.existing_refs_escrow if account_type == "escrow" else self.existing_refs_corporate
        cutoff_date = self.latest_date_escrow if account_type == "escrow" else self.latest_date_corporate
        new_count = 0

        for tx in parser.transactions:
            ref = tx.reference.strip()

            # DATE FILTER: only consider transactions on or after the latest date in master
            if cutoff_date:
                tx_date = self._parse_date(tx.date)
                if tx_date and tx_date < cutoff_date:
                    continue  # Skip older transactions, master sheet is correct till cutoff

            # Check if this reference already exists in the master sheet
            is_existing = False
            if ref in existing_refs:
                is_existing = True
            else:
                # Fuzzy: check if the ref's first 15 chars match any existing ref
                ref_prefix = ref[:15] if len(ref) > 15 else ref
                for existing in existing_refs:
                    if existing.startswith(ref_prefix) or ref.startswith(existing[:15]):
                        is_existing = True
                        break

            if is_existing:
                continue

            # This is a NEW transaction
            new_tx = NewTransaction(
                account=account_type,
                date=tx.date,
                value_date=tx.value_date,
                narration=tx.description[:200],
                reference=ref,
                debit=tx.debit,
                credit=tx.credit,
                balance=tx.balance,
            )

            # Match to unit using multiple methods
            self._match_new_transaction(new_tx, parser)

            self.new_transactions.append(new_tx)
            new_count += 1

        return new_count

    # =================================================================
    # STEP 3: Match new transactions to units
    # =================================================================

    def _collect_fuzzy_alternatives(self, normalized_name: str, chosen_unit: str, max_alts: int = 3) -> list:
        """Collect top fuzzy-match alternatives (excluding the chosen unit) for review UI."""
        _, _, all_candidates = self._fuzzy_name_match(normalized_name, return_all=True)
        # Filter out the chosen match and deduplicate by unit_no
        seen = {chosen_unit} if chosen_unit else set()
        alts = []
        for c in all_candidates:
            if c["unit_no"] not in seen:
                seen.add(c["unit_no"])
                alts.append(c)
            if len(alts) >= max_alts:
                break
        return alts

    def _match_new_transaction(self, tx: NewTransaction, parser: BankStatementParser):
        """Match a new transaction to a unit using all available methods.
        For low-confidence or fuzzy matches, also populates alternative_matches
        with the top 3 runner-up candidates.

        FIX 8: After collecting all candidates (substring, fuzzy, narration),
        a candidate matching more name parts always beats one matching fewer parts."""

        # Method 1: Unit number in narration
        unit, conf = parser.extract_unit_from_description(tx.narration)
        if unit and conf >= 0.80:
            tx.unit_no = unit
            tx.match_confidence = conf
            tx.match_method = "unit_in_narration"
            # Look up name from knowledge base
            names_list = self.kb_unit_to_name.get(unit, [])
            tx.account_name = names_list[0] if names_list else ""
            return

        # Collect all candidates from multiple methods, then pick the best (FIX 8)
        candidates = []  # list of (unit, conf, method, name, matched_parts, total_parts)

        # Method 2: Name in narration -> match to known unit
        name = parser.extract_name_from_description(tx.narration)
        if name and len(name) > 3:
            normalized = self._normalize_name(name)
            norm_parts = [p for p in normalized.split() if len(p) > 2]

            # Exact KB match
            if normalized in self.kb_name_to_unit:
                kb_unit = self.kb_name_to_unit[normalized][0]
                candidates.append((kb_unit, 0.85, "name_from_kb", name, len(norm_parts), len(norm_parts)))

            # FIX 1: Best-match substring (collect ALL, pick best overlap)
            for kb_name, kb_units in self.kb_name_to_unit.items():
                if len(normalized) > 5 and len(kb_name) > 5:
                    if normalized in kb_name or kb_name in normalized:
                        overlap = min(len(normalized), len(kb_name))
                        total = max(len(normalized), len(kb_name))
                        ratio = overlap / total if total > 0 else 0
                        kb_parts = [p for p in kb_name.split() if len(p) > 2]
                        matched_p = sum(1 for p in norm_parts if p in kb_parts)
                        candidates.append((kb_units[0], 0.82 * ratio / 0.5 if ratio < 0.5 else 0.82,
                                           "name_substring", name, matched_p, max(len(norm_parts), len(kb_parts))))

            # Fuzzy name match (improved with Arabic variants) - FIX 4: threshold 0.55
            best_unit, best_score, all_candidates = self._fuzzy_name_match(normalized, return_all=True)
            if best_unit and best_score >= 0.55:
                # Count matched parts for this candidate
                best_kb_name = ""
                for kn in self.kb_name_to_unit:
                    if self.kb_name_to_unit[kn][0] == best_unit:
                        best_kb_name = kn
                        break
                kb_parts = [p for p in best_kb_name.split() if len(p) > 2] if best_kb_name else []
                matched_p = sum(1 for p in norm_parts if any(
                    p == kp or self._is_variant(p, kp) for kp in kb_parts)) if kb_parts else 0
                candidates.append((best_unit, round(min(0.80, best_score * 0.85), 2),
                                   "name_fuzzy", name, matched_p, max(len(norm_parts), len(kb_parts)) if kb_parts else len(norm_parts)))

            # Store extracted name even if no match yet
            tx.account_name = name
            # For unmatched, still provide fuzzy alternatives if any exist
            if all_candidates:
                seen = set()
                alts = []
                for c in all_candidates:
                    if c["unit_no"] not in seen:
                        seen.add(c["unit_no"])
                        alts.append(c)
                    if len(alts) >= 3:
                        break
                tx.alternative_matches = alts

        # Method 2b: Direct narration scan (FIX 2: scored approach)
        narr_upper = tx.narration.upper()
        best_narr_unit = ""
        best_narr_score = 0.0
        best_narr_parts_matched = 0
        best_narr_total_parts = 0
        best_narr_kb_name = ""
        for kb_name, kb_units in self.kb_name_to_unit.items():
            parts = [p for p in kb_name.split() if len(p) > 2]
            if len(parts) < 2:
                continue
            last = parts[-1]
            if last in narr_upper and any(p in narr_upper for p in parts[:-1]):
                matching_parts = sum(1 for p in parts if p in narr_upper)
                score = matching_parts / len(parts)
                if score > best_narr_score:
                    best_narr_score = score
                    best_narr_unit = kb_units[0]
                    best_narr_kb_name = kb_name
                    best_narr_parts_matched = matching_parts
                    best_narr_total_parts = len(parts)
        if best_narr_unit:
            candidates.append((best_narr_unit, 0.75, "name_in_narration",
                               name or "", best_narr_parts_matched, best_narr_total_parts))

        # FIX 8: Pick the best candidate — more matched parts always wins
        if candidates:
            # Sort by (matched_parts desc, confidence desc) so full-name beats partial-name
            candidates.sort(key=lambda c: (c[4], c[1]), reverse=True)
            best = candidates[0]
            tx.unit_no = best[0]
            tx.match_confidence = best[1]
            tx.match_method = best[2]
            if best[3]:
                tx.account_name = best[3]
            if not tx.account_name:
                names_list = self.kb_unit_to_name.get(tx.unit_no, [])
                tx.account_name = names_list[0] if names_list else ""
            # Collect alternatives
            if name and len(name) > 3:
                tx.alternative_matches = self._collect_fuzzy_alternatives(
                    self._normalize_name(name), tx.unit_no)
            return

        # Method 3: IBAN/account matching from narration
        iban = parser.extract_iban_from_description(tx.narration)
        if iban:
            # Check if we've seen this IBAN before with a unit
            for ref, known_unit in self.kb_ref_to_unit.items():
                # This is a simplistic check; in practice you'd track IBANs
                pass

        # No match found
        tx.match_method = "unmatched"
        tx.match_confidence = 0.0

    @staticmethod
    def _is_variant(name_a: str, name_b: str) -> bool:
        """Check if two name tokens are Arabic transliteration variants of each other."""
        for canonical, variants in NAME_VARIANTS.items():
            all_forms = {canonical} | set(variants)
            if name_a in all_forms and name_b in all_forms:
                return True
        return False

    def _fuzzy_name_match(self, normalized_name: str, return_all: bool = False):
        """Fuzzy match a name against the knowledge base.
        Uses exact token matching first (most reliable), then variant expansion.
        Prioritizes matches where the LAST NAME (surname) matches exactly.

        If return_all=True, returns (best_unit, best_score, all_candidates) where
        all_candidates is a list of {unit_no, account_name, confidence, method} dicts
        for every candidate scoring above 0.3, sorted descending by score."""
        parts = normalized_name.split()
        if len(parts) < 2:
            return ("", 0.0, []) if return_all else ("", 0.0)

        significant_parts = [p for p in parts if len(p) > 2]
        if not significant_parts:
            return ("", 0.0, []) if return_all else ("", 0.0)

        best_unit = ""
        best_score = 0.0
        all_candidates = []  # collect all candidates above threshold

        for kb_name, kb_units in self.kb_name_to_unit.items():
            kb_unit = kb_units[0]  # FIX 7: kb_name_to_unit values are now lists
            kb_parts = [k for k in kb_name.split() if len(k) > 2]
            if not kb_parts:
                continue

            # Phase 1: EXACT token match (no variant expansion)
            exact_matches = sum(1 for p in significant_parts if p in kb_parts)
            exact_score = exact_matches / max(len(significant_parts), len(kb_parts))

            # Phase 2: Last name (surname) must match for high confidence
            last_name = significant_parts[-1] if significant_parts else ""
            last_name_matches = last_name in kb_parts

            # Phase 3: Variant-aware matching (only for first/middle names)
            variant_matches = 0
            for p in significant_parts:
                if p in kb_parts:
                    variant_matches += 1
                    continue
                # Check Arabic name variants
                for canonical, variants in NAME_VARIANTS.items():
                    all_forms = {canonical} | set(variants)
                    if p in all_forms and any(kp in all_forms for kp in kb_parts):
                        variant_matches += 1
                        break

            variant_score = variant_matches / max(len(significant_parts), len(kb_parts))

            # Final score: prioritize exact matches, require surname match for high confidence
            if last_name_matches and exact_matches >= 2:
                # Strong match: surname + 2+ other parts match exactly
                score = min(1.0, exact_score + 0.1)
            elif last_name_matches:
                # Surname matches but fewer exact tokens
                score = min(0.85, variant_score)
            else:
                # FIX 3: Surname doesn't match — much lower confidence (was 0.5/0.6)
                score = min(0.3, variant_score * 0.4)

            # Collect all candidates above 0.3 for alternative display
            if return_all and score >= 0.3:
                names_list = self.kb_unit_to_name.get(kb_unit, [])
                all_candidates.append({
                    "unit_no": kb_unit,
                    "account_name": names_list[0] if names_list else "",
                    "confidence": round(min(1.0, score), 3),
                    "method": "name_fuzzy",
                })

            # FIX 4: Raised threshold from 0.4 to 0.55
            if score > best_score and score >= 0.55:
                best_score = score
                best_unit = kb_unit

        if return_all:
            # Sort candidates by confidence descending
            all_candidates.sort(key=lambda c: c["confidence"], reverse=True)
            return (best_unit, min(best_score, 1.0), all_candidates)

        return (best_unit, min(best_score, 1.0))

    # =================================================================
    # STEP 4: Update master sheet with new transactions
    # =================================================================

    def update_master_sheet(self, master_path: str, output_path: str) -> dict:
        """Append new transactions to the master sheet and save."""
        wb = openpyxl.load_workbook(master_path)

        escrow_added = 0
        corporate_added = 0

        # Styles for new rows
        new_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        matched_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        unmatched_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        for tx in self.new_transactions:
            if tx.account == "escrow" and "Updated Sheet_Escrow Account" in wb.sheetnames:
                ws = wb["Updated Sheet_Escrow Account"]
                next_row = ws.max_row + 1
                ws.cell(next_row, 1, tx.date)
                ws.cell(next_row, 2, tx.value_date)
                ws.cell(next_row, 3, tx.narration)
                ws.cell(next_row, 4, tx.reference)
                ws.cell(next_row, 5, tx.debit if tx.debit else 0)
                ws.cell(next_row, 6, tx.credit if tx.credit else 0)
                # Running Balance formula: =G{prev}-E{row}+F{row}
                prev_row = next_row - 1
                ws.cell(next_row, 7, f"=G{prev_row}-E{next_row}+F{next_row}")
                ws.cell(next_row, 8, f"{tx.unit_no}" if tx.unit_no else "UNMATCHED")
                ws.cell(next_row, 9, tx.account_name)
                ws.cell(next_row, 10, "")  # Receipt
                ws.cell(next_row, 11, "NEW" if tx.unit_no else "PENDING")

                fill = matched_fill if tx.unit_no else unmatched_fill
                for c in range(1, 12):
                    ws.cell(next_row, c).fill = fill

                escrow_added += 1

            elif tx.account == "corporate" and "Updated Sheet_Corporate" in wb.sheetnames:
                ws = wb["Updated Sheet_Corporate"]
                next_row = ws.max_row + 1
                ws.cell(next_row, 1, tx.date)
                ws.cell(next_row, 2, tx.narration)
                ws.cell(next_row, 3, tx.reference)
                ws.cell(next_row, 4, tx.debit if tx.debit else 0)
                ws.cell(next_row, 5, tx.credit if tx.credit else 0)
                # Running Balance formula: =F{prev}-D{row}+E{row}
                prev_row = next_row - 1
                ws.cell(next_row, 6, f"=F{prev_row}-D{next_row}+E{next_row}")
                ws.cell(next_row, 7, f"{tx.unit_no}" if tx.unit_no else "UNMATCHED")
                ws.cell(next_row, 8, tx.account_name)

                fill = matched_fill if tx.unit_no else unmatched_fill
                for c in range(1, 9):
                    ws.cell(next_row, c).fill = fill

                corporate_added += 1

        # Also update previously unreconciled rows that we matched
        updated_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # light blue
        unreconciled_updated = 0

        for ur in self._unreconciled_rows:
            if not ur.get("unit_no"):
                continue
            sheet = ur["sheet"]
            row = ur["row"]
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                unit_col = 8 if sheet.startswith("Updated Sheet_E") else 7
                name_col = 9 if sheet.startswith("Updated Sheet_E") else 8

                ws.cell(row, unit_col, str(ur["unit_no"]))
                ws.cell(row, name_col, ur.get("account_name", ""))

                for c in range(1, name_col + 1):
                    ws.cell(row, c).fill = updated_fill

                unreconciled_updated += 1

        wb.save(output_path)
        return {
            "escrow_added": escrow_added,
            "corporate_added": corporate_added,
            "total_added": escrow_added + corporate_added,
            "unreconciled_updated": unreconciled_updated,
            "output_path": output_path,
        }

    # =================================================================
    # STEP 5: Generate receipts for matched credits
    # =================================================================

    def generate_receipts(self, output_dir: str) -> list[dict]:
        """Generate receipt PDFs for all matched new credit transactions."""
        os.makedirs(output_dir, exist_ok=True)
        self.receipts_generated = []

        for tx in self.new_transactions:
            if not tx.unit_no or tx.credit <= 0:
                continue

            receipt_no = f"REC-{tx.date.replace('-', '')}-{tx.unit_no}-{tx.reference[:8]}"
            filename = f"Receipt_{tx.unit_no}_{tx.date.replace('-', '')}_{receipt_no[-6:]}.pdf"
            filepath = os.path.join(output_dir, filename)

            try:
                generate_receipt(
                    output_path=filepath,
                    receipt_no=receipt_no,
                    unit_no=tx.unit_no,
                    client_name=tx.account_name or "Valued Client",
                    payment_date=tx.date,
                    amount=tx.credit,
                    installment_type=self._guess_installment_type(tx),
                    payment_method="Bank Transfer",
                    reference=tx.reference,
                    narration=tx.narration,
                )

                tx.receipt_generated = True
                tx.receipt_path = filepath

                self.receipts_generated.append({
                    "unit_no": tx.unit_no,
                    "client_name": tx.account_name,
                    "amount": tx.credit,
                    "receipt_no": receipt_no,
                    "filepath": filepath,
                    "account": tx.account,
                })
            except Exception as e:
                print(f"  Error generating receipt for Unit {tx.unit_no}: {e}")

        return self.receipts_generated

    def _guess_installment_type(self, tx: NewTransaction) -> str:
        """Guess the installment type from the narration."""
        narr = tx.narration.upper()
        if "DLD" in narr or "REGISTRATION" in narr:
            return "DLD Registration Fee"
        if "BOOKING" in narr or "DOWN PAYMENT" in narr:
            return "Booking / Down Payment"
        if "HANDOVER" in narr:
            return "On Handover"
        if re.search(r'Q[1-8]', narr):
            m = re.search(r'(Q[1-8])', narr)
            return f"{m.group(1)} Installment"
        if tx.account == "corporate":
            return "Commission / DLD Fee"
        return "Installment Payment"

    # =================================================================
    # STEP 6: Claude AI Matching for remaining unmatched
    # =================================================================

    def run_ai_matching(self, api_key: str = None) -> list[dict]:
        """
        Use Claude to analyze unmatched transactions and suggest matches.
        Returns list of AI match suggestions with reasoning.
        """
        from config.settings import ANTHROPIC_API_KEY, CLAUDE_MODEL, CLAUDE_MAX_TOKENS, AI_MATCHING_PROMPT

        key = api_key or os.environ.get("ANTHROPIC_API_KEY", ANTHROPIC_API_KEY)
        if not key or not HAS_ANTHROPIC:
            return []

        unmatched = [t for t in self.new_transactions if not t.unit_no and t.credit > 0]
        if not unmatched:
            return []

        # Build context for Claude
        kb_summary = []
        for unit, names in sorted(self.kb_unit_to_name.items()):
            kb_summary.append(f"  Unit {unit}: {', '.join(names)}")

        tx_list = []
        for i, tx in enumerate(unmatched):
            tx_list.append(
                f"  [{i}] Date: {tx.date} | Amount: AED {tx.credit:,.2f} | "
                f"Narration: {tx.narration} | Ref: {tx.reference} | "
                f"Extracted name: {tx.account_name or 'none'}"
            )

        prompt = f"""Analyze these unmatched bank transactions and match them to units.

KNOWLEDGE BASE — Unit to Client Mappings:
{chr(10).join(kb_summary[:150])}

UNMATCHED TRANSACTIONS:
{chr(10).join(tx_list)}

For each transaction, respond with a JSON array where each element has:
- "index": transaction index from the list above
- "unit_no": matched unit number (string) or "" if no match
- "confidence": 0.0 to 1.0
- "reasoning": brief explanation
- "match_method": "ai_name_match" or "ai_amount_match" or "ai_pattern" or "ai_no_match"

Only return the JSON array, no other text."""

        try:
            client = anthropic.Anthropic(api_key=key)
            response = client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=CLAUDE_MAX_TOKENS,
                system=AI_MATCHING_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )

            # Parse response
            text = response.content[0].text.strip()
            # Extract JSON from response (handle markdown code blocks)
            if "```" in text:
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
                text = text.strip()

            suggestions = json.loads(text)
            ai_results = []

            for suggestion in suggestions:
                idx = suggestion.get("index", -1)
                unit_no = suggestion.get("unit_no", "")
                confidence = float(suggestion.get("confidence", 0))
                reasoning = suggestion.get("reasoning", "")
                method = suggestion.get("match_method", "ai_match")

                if 0 <= idx < len(unmatched) and unit_no and confidence >= 0.6:
                    tx = unmatched[idx]
                    tx.unit_no = unit_no
                    tx.match_confidence = confidence
                    tx.match_method = method
                    # Look up name from KB if not already set
                    if not tx.account_name:
                        names_list = self.kb_unit_to_name.get(unit_no, [])
                        tx.account_name = names_list[0] if names_list else ""

                ai_results.append({
                    "index": idx,
                    "narration": unmatched[idx].narration if 0 <= idx < len(unmatched) else "",
                    "amount": unmatched[idx].credit if 0 <= idx < len(unmatched) else 0,
                    "unit_no": unit_no,
                    "confidence": confidence,
                    "reasoning": reasoning,
                    "method": method,
                    "applied": bool(unit_no and confidence >= 0.6),
                })

            return ai_results

        except Exception as e:
            print(f"  AI matching error: {e}")
            return [{"error": str(e)}]

    # =================================================================
    # Salesforce Integration
    # =================================================================

    def sync_with_salesforce(self, sf_service=None, project_name: str = None) -> list[dict]:
        """
        For each matched transaction, look up the unit in Salesforce,
        identify which installment(s) the payment covers, and prepare
        SF actions (update payment status, create receipt, send email).
        All actions respect SF_WRITE_MODE and EMAIL_LIVE_MODE flags.
        """
        if sf_service is None:
            return []

        sf_actions = []
        matched = [t for t in self.new_transactions if t.unit_no and t.credit > 0]

        for tx in matched:
            try:
                # Identify installments covered by this payment
                installments = sf_service.identify_installments_for_amount(
                    unit_name=tx.unit_no,
                    bank_amount=tx.credit,
                    project_name=project_name,
                )

                if not installments:
                    sf_actions.append({
                        "unit": tx.unit_no,
                        "amount": tx.credit,
                        "action": "no_sf_match",
                        "message": f"No matching installment found in SF for AED {tx.credit:,.2f}",
                        "results": [],
                    })
                    continue

                tx_actions = []
                for inst in installments:
                    payment = inst["payment"]
                    allocated = inst["allocated_amount"]
                    match_type = inst["match_type"]

                    # Update payment amount_paid
                    new_paid = payment.amount_paid + allocated
                    update_result = sf_service.update_payment_amount_paid(
                        booking_payment_id=payment.sf_id,
                        new_amount_paid=new_paid,
                        notes=f"Bank ref: {tx.reference} | {tx.narration[:50]}",
                    )
                    tx_actions.append(update_result.to_dict())

                    # Create receipt
                    receipt_result = sf_service.create_receipt_record(
                        inventory_id=payment.inventory_id,
                        amount=allocated,
                        payment_type=payment.payment_type,
                        payment_date=tx.date,
                        reference=tx.reference,
                    )
                    tx_actions.append(receipt_result.to_dict())

                sf_actions.append({
                    "unit": tx.unit_no,
                    "amount": tx.credit,
                    "action": "sf_sync",
                    "installments_matched": len(installments),
                    "match_type": installments[0]["match_type"] if installments else "",
                    "message": f"Matched {len(installments)} installment(s)",
                    "results": tx_actions,
                })

            except Exception as e:
                sf_actions.append({
                    "unit": tx.unit_no,
                    "amount": tx.credit,
                    "action": "error",
                    "message": str(e),
                    "results": [],
                })

        return sf_actions

    # =================================================================
    # SUMMARY
    # =================================================================

    def get_summary(self) -> dict:
        """Get a summary of the daily reconciliation."""
        matched = [t for t in self.new_transactions if t.unit_no]
        unmatched = [t for t in self.new_transactions if not t.unit_no]

        # Break down by match method
        by_method = {}
        for t in matched:
            m = t.match_method or "unknown"
            by_method[m] = by_method.get(m, 0) + 1

        # Unreconciled rows from master sheet
        ur_matched = [u for u in self._unreconciled_rows if u.get("unit_no")]
        ur_still = [u for u in self._unreconciled_rows if not u.get("unit_no")]

        return {
            "total_new_transactions": len(self.new_transactions),
            "matched": len(matched),
            "unmatched": len(unmatched),
            "receipts_generated": len(self.receipts_generated),
            "by_account": {
                "escrow": len([t for t in self.new_transactions if t.account == "escrow"]),
                "corporate": len([t for t in self.new_transactions if t.account == "corporate"]),
            },
            "by_method": by_method,
            "matched_details": [t.to_dict() for t in matched],
            "unmatched_details": [t.to_dict() for t in unmatched],
            "receipts": self.receipts_generated,
            "unreconciled": {
                "total_found": len(self._unreconciled_rows),
                "now_matched": len(ur_matched),
                "still_unmatched": len(ur_still),
                "matched_details": ur_matched,
                "unmatched_details": ur_still,
            },
            "needs_receipt": self._needs_receipt if hasattr(self, '_needs_receipt') else [],
        }

    @staticmethod
    def _normalize_name(name: str) -> str:
        """Normalize a name for matching."""
        n = name.upper().strip()
        n = re.sub(r'\b([A-Z])\s+(?=[A-Z]{2,})', r'\1', n)
        n = re.sub(r'\s+', ' ', n)
        return n

    @staticmethod
    def _parse_date(val) -> datetime:
        """Parse a date value from Excel (string or datetime object).
        Handles DD-MM-YYYY strings and Excel datetime objects where
        openpyxl may have swapped month/day."""
        if not val:
            return None

        today = datetime.now()

        if isinstance(val, datetime):
            # If the date is in the future, month and day may be swapped
            # e.g., datetime(2026, 7, 3) should be March 7 not July 3
            if val > today and val.day <= 12:
                try:
                    swapped = val.replace(month=val.day, day=val.month)
                    if swapped <= today:
                        return swapped
                except ValueError:
                    pass
            return val

        s = str(val).strip()
        # Try DD-MM-YYYY first (most common in ENBD statements)
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y"):
            try:
                return datetime.strptime(s.split(" ")[0] if " " in s else s, fmt)
            except ValueError:
                continue
        # Then ISO format
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s.split(" ")[0] if " " in s else s, fmt)
            except ValueError:
                continue
        return None
